# -*- coding: utf-8 -*-
"""
Created on Sat Jan  6 18:28:30 2018

@author: CheukYan

@TO BE PLACED IN THE SAME LOCATION AS THE WORKBOOK!
"""
import os
import sys
import openpyxl
import datetime

bookname = 'AIDE MEMOIRE.xlsx' #put .py file in the location of the .xlsx and enter its name here

#Date
now = datetime.datetime.now() #acquire date information
date = str(now.strftime("%Y-%m-%d %H:%M")) #formatting the date information

#Open Workbook
os.getcwd() #current working directory. It will tell the current working directory for python, the result may be like that, as it is in my interpreter.
book = openpyxl.load_workbook(bookname)

#Get Sheet Names
sheetnames = book.get_sheet_names() #acquire sheet names in an array

while True:
#Select Sheet
    print('AIDE MEMOIRE LOG INPUT ASSISTANT')
    sheetnum = int(input('ISSUES=1, MISC=2, QUIT=0 : '))
    
    if sheetnum == 0:
        sys.exit()
    if sheetnum == 1:
        sheet = book.get_sheet_by_name(sheetnames[sheetnum-1]) #load sheet 
        rownum = str(sheet.max_row + 1) #find the next empty row
        #INPUTS
        issue = input('What is the issue? : ')
        solution = input('How did you solve it? : ')
        #PRINT STUFF
        sheet['A' + rownum] = date
        sheet['B' + rownum] = issue
        sheet['C' + rownum] = issue
        #SAVE TO SHEET & QUIT
        book.save(bookname)
        continue#exit()
    if sheetnum == 2:
        sheet = book.get_sheet_by_name(sheetnames[sheetnum-1]) #load sheet 
        rownum = str(sheet.max_row + 1) #find the next empty row
        #INPUTS
        misc = input('What is it? : ')
        #PRINT STUFF
        sheet['A' + rownum] = date
        sheet['B'+ rownum] = misc
        #SAVE TO SHEET & QUIT
        book.save(bookname)
        continue#exit()
    else:
        print('This is not a valid input')
        continue